from __future__ import annotations

import io
import json
import logging
import os
import re
from typing import Any

from fastapi import HTTPException
from google.genai import types
from openpyxl import load_workbook

from API_RAG_NEW.concurrency import acquire_llm_slot
from API_RAG_NEW.config import GEMINI_MODEL, get_gemini_api_key
from llms.onlinellms import OnlineLLMs

logger = logging.getLogger(__name__)

MULTIMODAL_MIME_TYPES = {
    "application/pdf",
    "image/jpeg",
    "image/jpg",
    "image/png",
    "image/webp",
    "image/gif",
    "image/tiff",
}

SPREADSHEET_EXTENSIONS = {".xlsx"}

KIND_FIELD_HINTS: dict[str, str] = {
    "electricity_bill": "supplier, period_start, period_end, kwh_total, amount_vnd, tax_vnd, meter_number",
    "fuel_receipt": "supplier, date, fuel_type, liters, unit_price, amount_vnd",
    "material_invoice": "supplier, date, items, total_weight_kg, amount_vnd, currency",
    "warehouse_receipt": "product_name, quantity, unit, issued_date, warehouse_code",
    "bill_of_lading": "carrier, origin_port, destination_port, vessel_name, container_count, weight_kg",
    "air_waybill": "airline, origin_airport, destination_airport, awb_number, weight_kg",
    "supplier_certificate": "supplier_name, standard, issued_date, expiry_date, issuing_body",
    "supplier_declaration": "supplier_name, declaration_date, scope, signatory",
    "bom": "product_name, materials_list, quantities, units",
    "export_invoice": "buyer, date, items, quantities, unit_prices, total_amount, currency",
    "packing_list": "shipper, consignee, packages_count, total_weight_kg, total_volume_m3",
    "logistics_invoice": "carrier, service, origin, destination, weight_kg, amount_vnd",
}

DEFAULT_HINT = "organization_name, date, numeric_values_with_units, total_amount"

JSON_OBJECT_PATTERN = re.compile(r"\{[\s\S]*?\}", re.DOTALL)
JSON_GREEDY_PATTERN = re.compile(r"\{[\s\S]*\}", re.DOTALL)


def _build_llm() -> OnlineLLMs:
    api_key = get_gemini_api_key()
    if not api_key:
        raise HTTPException(status_code=400, detail="GEMINI_API_KEY not configured")
    return OnlineLLMs(name="gemini", api_key=api_key, model_version=GEMINI_MODEL)


def _build_prompt(kind: str, language: str) -> str:
    hint = KIND_FIELD_HINTS.get(kind, DEFAULT_HINT)
    return (
        "Bạn là chuyên gia kiểm toán carbon (GHG Protocol / CBAM).\n"
        f"Loại chứng từ: {kind}.\n"
        f"Trích xuất các trường sau (ưu tiên): {hint}.\n"
        "Trả về JSON object thuần túy (chỉ JSON, không markdown, không giải thích).\n"
        "Key: snake_case tiếng Anh. Value: giá trị tìm thấy hoặc null.\n"
        f"Ngôn ngữ nhận dạng: {language}."
    )


def _parse_json(text: str) -> dict[str, Any]:
    # Try greedy match first (captures full nested object)
    match = JSON_GREEDY_PATTERN.search(text or "")
    if not match:
        return {}
    try:
        payload = json.loads(match.group(0))
        return payload if isinstance(payload, dict) else {}
    except json.JSONDecodeError:
        pass
    # Fallback: minimal match
    match = JSON_OBJECT_PATTERN.search(text or "")
    if not match:
        return {}
    try:
        payload = json.loads(match.group(0))
        return payload if isinstance(payload, dict) else {}
    except json.JSONDecodeError:
        return {}


def _extract_xlsx_text(raw_bytes: bytes) -> str:
    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    try:
        blocks: list[str] = []
        for sheet in workbook.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [
                    str(cell).strip()
                    for cell in row
                    if cell is not None and str(cell).strip()
                ]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                blocks.append(f"[Sheet: {sheet.title}]\n" + "\n".join(rows))
        return "\n\n".join(blocks)
    finally:
        workbook.close()


def extract_document_fields(
    filename: str,
    raw_bytes: bytes,
    mime_type: str | None,
    kind: str,
    language: str = "vi",
) -> dict[str, Any]:
    llm = _build_llm()
    prompt_text = _build_prompt(kind, language)
    effective_mime = (mime_type or "").lower().split(";")[0].strip()
    extension = os.path.splitext(filename)[1].casefold()

    logger.info(
        "[extract] start filename=%r kind=%s mime=%r ext=%s bytes=%d model=%s",
        filename, kind, effective_mime, extension, len(raw_bytes or b""),
        llm.model_version,
    )

    try:
        if effective_mime in MULTIMODAL_MIME_TYPES and raw_bytes:
            branch = "multimodal"
            contents = [
                types.Part.from_bytes(data=raw_bytes, mime_type=effective_mime),
                types.Part.from_text(text=prompt_text),
            ]
            with acquire_llm_slot():
                response = llm.client.models.generate_content(
                    model=llm.model_version,
                    contents=contents,
                )
            raw_text = response.text or ""
        elif extension in SPREADSHEET_EXTENSIONS and raw_bytes:
            branch = "xlsx"
            file_text = _extract_xlsx_text(raw_bytes)[:8000]
            logger.info("[extract] xlsx text_chars=%d", len(file_text))
            combined = f"Nội dung file ({filename}):\n{file_text}\n\n{prompt_text}"
            raw_text = llm.generate_content(combined)
        else:
            branch = "text"
            try:
                file_text = raw_bytes.decode("utf-8", errors="replace")[:8000]
            except Exception:
                file_text = ""
            logger.info(
                "[extract] text branch mime=%r ext=%s decoded_chars=%d "
                "(no multimodal/xlsx match — binary files land here and usually yield nothing)",
                effective_mime, extension, len(file_text),
            )
            combined = f"Nội dung file ({filename}):\n{file_text}\n\n{prompt_text}"
            raw_text = llm.generate_content(combined)

        fields = _parse_json(raw_text)
        raw_len = len(raw_text or "")
        if not raw_text.strip():
            logger.warning(
                "[extract] EMPTY model output filename=%r branch=%s — Gemini returned "
                "nothing (check API key/quota/model name/safety-block)",
                filename, branch,
            )
        elif not fields:
            logger.warning(
                "[extract] NO FIELDS parsed filename=%r branch=%s raw_len=%d "
                "raw_preview=%r — model replied but no JSON object found",
                filename, branch, raw_len, (raw_text or "")[:300],
            )
        else:
            logger.info(
                "[extract] ok filename=%r branch=%s raw_len=%d fields=%d keys=%s",
                filename, branch, raw_len, len(fields), list(fields.keys())[:20],
            )
        return fields

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(
            "[extract] FAILED filename=%r kind=%s mime=%r ext=%s: %s",
            filename, kind, effective_mime, extension, exc,
        )
        raise HTTPException(
            status_code=502, detail=f"AI extraction failed: {exc}"
        ) from exc
